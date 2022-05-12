import time
import openpyxl
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.select import Select
from webdriver_manager.chrome import ChromeDriverManager
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://www.barcodelookup.com/")
driver.maximize_window()
time.sleep(2)
list_barcode=[]
barcode=openpyxl.load_workbook("C:\\Users\\Dell\\Desktop\\barcode.xlsx")
barc1=barcode.active
max_rows=barc1.max_row
print(max_rows)
max_cols = barc1.max_column

for i in range(2,max_rows+1):
    for j in range(1,max_cols+1):
        #print(barc1.cell(row=i, column=j).value)
        list_barcode.append(barc1.cell(row=i, column=j).value)

print(list_barcode)



#Taking each barcode and search
mlist=[]
for barcode in list_barcode:
    #print(barcode,"inside loop")
    #barcode_input = driver.find_element(By.XPATH, "//input[@name='search-input']")
    barcode_input = driver.find_element(By.NAME, "search-input")
    barcode_input.click()
    barcode_input.send_keys(barcode)


    driver.find_element(By.XPATH,"//button[@class='btn-search']").click()
   # continue
   # if(barcode !=''):
    try:
        manufacturer = driver.find_element(By.XPATH,"//*[@id='product']/section[2]/div[1]/div/div/div[2]/div[5]/span")
        #print(manufacturer.text)
        mlist.append(manufacturer.text)
    except NoSuchElementException:
        #print("Invalid Bar Code")
        mlist.append("Invalid Bar Code")

    time.sleep(10)
print("Manufacturer list : ",mlist)
#for l in mlist:
   # barc1["B2:B6"]=l
    #barc1.save(filename="barcode.xlsx")

print("done")
barc1.append(mlist)
