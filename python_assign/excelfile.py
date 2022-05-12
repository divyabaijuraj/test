'''
import openpyxl

path = "C:\\Users\\Dell\\Desktop\\t1.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
         print(sheet.cell(i,j).value,end=" ")
    print("\n")
'''
import openpyxl

path = "C:\\Users\\Dell\\Desktop\\t1.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
         sheet.cell(i,j).value='Welcome'
    print("\n")
workbook.save(path)