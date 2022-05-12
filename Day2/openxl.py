from openpyxl import load_workbook


class XL:
    def read_data(self, file_name, ws_name):
        print("**************")
        # Load the workbook
        work_book_obj = load_workbook(filename=file_name)
        print(work_book_obj)
        #Assign the work sheet
        work_sheet_obj = work_book_obj[ws_name]

        xl_data = []

        # Fetching max no of row
        r_count = work_sheet_obj.max_row

        # Fetching max no of row
        c_count = work_sheet_obj.max_column

        row_data = []

        for i in range(1, r_count+1):
            for j in range(1, c_count+1):
                row_data.append(work_sheet_obj.cell(row=i, column=j).value)

            xl_data.append(row_data)
            row_data = []
        return xl_data

xl_obj = XL()

print(xl_obj.read_data("C:\\Users\\Dell\\Desktop\\test.xlsx", "Sheet1"))
